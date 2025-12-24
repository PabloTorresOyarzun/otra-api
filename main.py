from fastapi import FastAPI, HTTPException, APIRouter, File, UploadFile, Security, Depends
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
import httpx
import os
import base64
import io
import fitz
import sys
import asyncio
import logging
import time
from concurrent.futures import ThreadPoolExecutor
from contextlib import contextmanager
from typing import Optional, Dict, List, Any
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
import pandas as pd
from PIL import Image
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

from quality import paso_1_analizar_documento, paso_2_corregir_rotacion
from clasificacion import clasificar_documento_completo, segmentar_pdf
from config import get_settings, calcular_timeout_excel, calcular_timeout_calidad, get_valid_api_tokens
from token_manager import token_manager
from modelos import modelos_router


@contextmanager
def suprimir_prints():
    """Suprime temporalmente la salida a stdout."""
    original_stdout = sys.stdout
    sys.stdout = io.StringIO()
    try:
        yield
    finally:
        sys.stdout = original_stdout


# --- Configuración de Fuente para PDFs ---
try:
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    FUENTE_PRINCIPAL = 'STSong-Light'
except Exception:
    FUENTE_PRINCIPAL = 'Helvetica'


# Configurar logger
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

settings = get_settings()
app = FastAPI(title="API Docs")

# Esquema de seguridad para autenticación con token
security = HTTPBearer()


def verify_api_token(credentials: HTTPAuthorizationCredentials = Security(security)) -> str:
    """
    Verifica que el token API proporcionado sea válido.

    Uso en Postman:
    - Header: Authorization
    - Value: Bearer <tu-token-aqui>
    """
    # Si es el ADMIN_TOKEN, permitir acceso (admin tiene acceso a todo)
    if settings.ADMIN_TOKEN and credentials.credentials == settings.ADMIN_TOKEN:
        return credentials.credentials

    # Verificar desde el gestor de tokens (modo persistente)
    if token_manager.is_valid_token(credentials.credentials):
        return credentials.credentials

    # Fallback: verificar desde API_TOKENS en .env (modo legacy)
    valid_tokens = get_valid_api_tokens()
    if valid_tokens and credentials.credentials in valid_tokens:
        return credentials.credentials

    raise HTTPException(
        status_code=401,
        detail="Token de autenticación inválido"
    )


def verify_admin_token(credentials: HTTPAuthorizationCredentials = Security(security)) -> str:
    """
    Verifica que el token de administrador sea válido.
    Solo permite acceso a endpoints de gestión de tokens.

    Uso en Postman:
    - Header: Authorization
    - Value: Bearer <admin-token>
    """
    if not settings.ADMIN_TOKEN:
        raise HTTPException(
            status_code=500,
            detail="ADMIN_TOKEN no configurado en el servidor"
        )

    if credentials.credentials != settings.ADMIN_TOKEN:
        raise HTTPException(
            status_code=403,
            detail="Se requiere token de administrador"
        )

    return credentials.credentials


BASE_URL = "https://backend.juanleon.cl"
ENDPOINT_DESPACHO = "/api/admin/despachos/{codigo}"
ENDPOINT_DOCUMENTOS = "/api/admin/documentos64/despacho/{codigo_visible}"

# Azure Document Intelligence Custom Models
AZURE_DI_ENDPOINT = "http://azure-di-custom:5000"
API_VERSION = "2022-08-31"  # Para gestión de modelos (build, list, delete)
API_VERSION_ANALYZE = "2024-11-30"  # Para análisis de documentos

# Mapeo de tipos de documento a model IDs de Azure DI
DOCUMENT_TYPE_TO_MODEL = {
    "FACTURA_COMERCIAL": "invoice",
    "DOCUMENTO_TRANSPORTE": "transport",
    "CERTIFICADO_ORIGEN": "origin",
    "LISTA_EMBALAJE": "packing-list",
    "CERTIFICADO_SANITARIO": "health",
    "POLIZA_SEGURO": "insurance",
    "UNKNOWN_DOCUMENT": None
}

# Almacenamiento en memoria
documentos_finales_sgd: Dict[str, List[Dict]] = {}
documentos_finales_individuales: Dict[str, List[Dict]] = {}

# Thread pool para operaciones CPU-bound (dinámico basado en CPU cores)
# max_workers = min(EXECUTOR_MAX_WORKERS, cpu_count * 4) para balancear rendimiento y recursos
executor = ThreadPoolExecutor(max_workers=min(settings.EXECUTOR_MAX_WORKERS, (os.cpu_count() or 1) * 4))
logger.info(f"ThreadPoolExecutor inicializado con max_workers={executor._max_workers}")

# Semaphore para limitar concurrencia de PDFs procesándose simultáneamente
MAX_CONCURRENT_PDFS = 10
pdf_semaphore = asyncio.Semaphore(MAX_CONCURRENT_PDFS)


class DocumentoSimplificado(BaseModel):
    nombre: str
    estado: str
    fecha_recepcion: str


class Usuarios(BaseModel):
    pedidor: Optional[List[str]] = None
    jefe_operaciones: Optional[List[str]] = None


class ConsultaResponse(BaseModel):
    codigo_despacho: str
    id_interno: str
    cliente: str
    estado: str
    tipo: str
    total_documentos: int
    documentos: List[DocumentoSimplificado]
    usuarios: Usuarios


class Alerta(BaseModel):
    pagina: int
    tipo: str
    descripcion: str


class DocumentoFinal(BaseModel):
    archivo_origen: str
    nombre_salida: str
    tipo: str
    paginas: List[int]
    alertas: Optional[List[Alerta]] = None
    datos_extraidos: Optional[Dict[str, Any]] = None


class ProcesamientoResponse(BaseModel):
    codigo_despacho: str
    cliente: str
    estado: str
    tipo: str
    total_documentos_segmentados: int
    documentos: List[DocumentoFinal]


class ProcesamientoIndividualResponse(BaseModel):
    archivo_origen: str
    total_documentos_segmentados: int
    documentos: List[DocumentoFinal]


class TokenInfo(BaseModel):
    id: str
    name: str
    masked_token: str
    created_at: str
    created_by: str
    last_used: Optional[str] = None
    is_active: bool = True


class TokenCreateRequest(BaseModel):
    name: str


class TokenCreateResponse(BaseModel):
    id: str
    token: str
    name: str
    created_at: str
    message: str


class TokenDeleteResponse(BaseModel):
    success: bool
    message: str


# Routers
sgd_router = APIRouter(prefix="/sgd", tags=["SGD"])
documentos_router = APIRouter(prefix="/documentos", tags=["Documentos"])
admin_router = APIRouter(prefix="/admin/tokens", tags=["Admin - Gestión de Tokens"])


def validar_pdf(file_bytes: bytes) -> bool:
    """Valida que los bytes sean un PDF válido."""
    if not file_bytes.startswith(b'%PDF'):
        return False
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        doc.close()
        return True
    except Exception:
        return False


def validar_excel(file_bytes: bytes, nombre_archivo: str) -> bool:
    """Valida que los bytes sean un archivo Excel válido."""
    extension = os.path.splitext(nombre_archivo.lower())[1]
    
    if extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        if not file_bytes.startswith(b'PK\x03\x04'):
            return False
    elif extension in ['.xls', '.xlsb']:
        if not file_bytes.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'):
            return False
    
    try:
        engine = 'xlrd' if extension == '.xls' else 'openpyxl'
        pd.read_excel(io.BytesIO(file_bytes), engine=engine, nrows=0)
        return True
    except Exception:
        return False


def validar_tamano_archivo(file_bytes: bytes) -> None:
    """Valida que el archivo no exceda el tamaño máximo."""
    file_size_mb = len(file_bytes) / (1024 * 1024)
    if file_size_mb > settings.MAX_FILE_SIZE_MB:
        raise HTTPException(
            status_code=413,
            detail=f"Archivo excede el tamaño máximo de {settings.MAX_FILE_SIZE_MB}MB"
        )


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception_type((httpx.TimeoutException, httpx.NetworkError))
)
async def consultar_despacho_detalle(codigo_interno: str, token: str) -> Optional[Dict]:
    if not token:
        return None
    
    url = f"{BASE_URL}{ENDPOINT_DESPACHO.format(codigo=str(codigo_interno))}"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {token}"
    }
    
    timeout = httpx.Timeout(
        connect=settings.TIMEOUT_CONNECT,
        read=settings.TIMEOUT_READ,
        write=settings.TIMEOUT_WRITE,
        pool=5.0
    )
    
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            response = await client.get(url, headers=headers)
            response.raise_for_status()
            
            if 'application/json' not in response.headers.get('Content-Type', ''):
                return None
            
            datos = response.json()
            return datos if isinstance(datos, dict) else None
    except (httpx.TimeoutException, httpx.NetworkError):
        raise
    except Exception:
        return None


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception_type((httpx.TimeoutException, httpx.NetworkError))
)
async def consultar_documentacion(codigo: str, token: str) -> Optional[List]:
    if not token:
        return None
    
    url = f"{BASE_URL}{ENDPOINT_DOCUMENTOS.format(codigo_visible=str(codigo))}"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {token}"
    }
    
    timeout = httpx.Timeout(
        connect=settings.TIMEOUT_CONNECT,
        read=settings.TIMEOUT_READ,
        write=settings.TIMEOUT_WRITE,
        pool=5.0
    )
    
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            response = await client.get(url, headers=headers)
            response.raise_for_status()
            
            if 'application/json' not in response.headers.get('Content-Type', ''):
                return None
            
            datos_json = response.json()
            if not isinstance(datos_json, dict):
                return None
            
            documentos = datos_json.get("data", [])
            return documentos if isinstance(documentos, list) else None
    except (httpx.TimeoutException, httpx.NetworkError):
        raise
    except Exception:
        return None


def es_archivo_excel(nombre_archivo: str) -> bool:
    """Verifica si un archivo es Excel basándose en su extensión."""
    extensiones_excel = ['.xls', '.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm']
    extension = os.path.splitext(nombre_archivo.lower())[1]
    return extension in extensiones_excel


def es_archivo_imagen(nombre_archivo: str) -> bool:
    """Verifica si un archivo es una imagen basándose en su extensión."""
    extensiones_imagen = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp']
    extension = os.path.splitext(nombre_archivo.lower())[1]
    return extension in extensiones_imagen


def validar_imagen(file_bytes: bytes, nombre_archivo: str) -> bool:
    """Valida que los bytes sean una imagen válida."""
    try:
        imagen = Image.open(io.BytesIO(file_bytes))
        imagen.verify()
        return True
    except Exception:
        return False


def formatear_celda(valor):
    """Formatea el valor de una celda para su representación en el PDF."""
    if pd.isna(valor):
        return ""

    if isinstance(valor, (float, int)):
        try:
            val_float = float(valor)
            if val_float.is_integer():
                return str(int(val_float))
            return str(val_float)
        except Exception:
            return str(valor)

    if isinstance(valor, pd.Timestamp):
        return valor.strftime('%Y-%m-%d')

    return str(valor)


def limpiar_dataframe(df):
    """Limpia el DataFrame eliminando columnas y filas vacías, detectando gaps."""
    df = df.fillna('')
    df = df.astype(str)
    df = df.replace('nan', '')

    densidades = []
    for col in df.columns:
        celdas_con_dato = (df[col].str.strip().str.len() > 0).sum()
        densidades.append(celdas_con_dato)

    umbral = 2

    hay_gap = False
    gap_actual = 0
    for d in densidades:
        if d < umbral:
            gap_actual += 1
            if gap_actual >= 3:
                hay_gap = True
                break
        else:
            gap_actual = 0

    if hay_gap:
        cols_con_datos_idx = [i for i, d in enumerate(densidades) if d >= umbral]

        if not cols_con_datos_idx:
            return pd.DataFrame()

        primer_col = cols_con_datos_idx[0]
        fin_bloque = cols_con_datos_idx[-1]

        for i in range(primer_col, len(densidades)):
            gap = 0
            for j in range(i, min(i + 5, len(densidades))):
                if densidades[j] < umbral:
                    gap += 1
                else:
                    break
            if gap >= 3:
                fin_bloque = i - 1
                break

        cols_seleccionadas = [i for i in range(primer_col, fin_bloque + 1) if densidades[i] >= umbral]
        if not cols_seleccionadas:
            return pd.DataFrame()
        df = df.iloc[:, cols_seleccionadas]
    else:
        cols_con_datos = []
        for col in df.columns:
            if df[col].str.strip().str.len().sum() > 0:
                cols_con_datos.append(col)
        df = df[cols_con_datos]

    df = df[df.apply(lambda row: row.str.strip().str.len().sum() > 0, axis=1)]

    return df.reset_index(drop=True)


def calcular_anchos_columnas_mejorado(df, disponible_width):
    """Calcula anchos de columnas óptimos basados en el contenido."""
    if df.empty:
        return []

    num_cols = len(df.columns)
    max_lens = []

    for col_idx in range(num_cols):
        col_data = df.iloc[:, col_idx].astype(str)
        longitudes = col_data.str.len()
        p90 = longitudes.quantile(0.9) if len(longitudes) > 0 else 0
        max_lens.append(max(p90, 5))

    total_chars = sum(max_lens)
    if total_chars == 0:
        total_chars = 1

    col_widths = [(l / total_chars) * disponible_width for l in max_lens]

    if num_cols <= 6:
        min_width = 0.8 * inch
        max_width = 5.0 * inch
    elif num_cols <= 10:
        min_width = 0.5 * inch
        max_width = 4.0 * inch
    else:
        min_width = 0.4 * inch
        max_width = 3.0 * inch

    col_widths = [max(min(w, max_width), min_width) for w in col_widths]

    current_total = sum(col_widths)
    if current_total > 0:
        factor = disponible_width / current_total
        col_widths = [w * factor for w in col_widths]

    return col_widths


def _convertir_imagen_a_pdf_sync(imagen_bytes: bytes, nombre_archivo: str) -> bytes:
    """Función síncrona interna para conversión de imagen a PDF."""
    try:
        imagen = Image.open(io.BytesIO(imagen_bytes))

        if imagen.mode == 'RGBA':
            fondo = Image.new('RGB', imagen.size, (255, 255, 255))
            fondo.paste(imagen, mask=imagen.split()[3])
            imagen = fondo
        elif imagen.mode != 'RGB':
            imagen = imagen.convert('RGB')

        ancho_img, alto_img = imagen.size
        ancho_pagina, alto_pagina = A4
        margen = 30

        ancho_disponible = ancho_pagina - (2 * margen)
        alto_disponible = alto_pagina - (2 * margen)

        escala = min(ancho_disponible / ancho_img, alto_disponible / alto_img)

        ancho_final = ancho_img * escala
        alto_final = alto_img * escala

        x = (ancho_pagina - ancho_final) / 2
        y = (alto_pagina - alto_final) / 2

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)

        img_reader = ImageReader(imagen)
        c.drawImage(img_reader, x, y, width=ancho_final, height=alto_final, preserveAspectRatio=True)
        c.save()

        return buffer.getvalue()

    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error al convertir imagen a PDF: {str(e)}"
        )


def _convertir_excel_a_pdf_sync(excel_bytes: bytes, nombre_archivo: str) -> bytes:
    """Función síncrona interna para conversión Excel a PDF con soporte avanzado."""
    try:
        extension = os.path.splitext(nombre_archivo.lower())[1]

        df_dict = None
        errores = []

        if extension in ['.xlsx', '.xlsm']:
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='openpyxl', header=None)
            except Exception as e1:
                errores.append(f"openpyxl: {str(e1)[:100]}")
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
                    df_dict = {}
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        data = []
                        for row in ws.iter_rows(values_only=True):
                            data.append(list(row))
                        if data:
                            df_dict[sheet_name] = pd.DataFrame(data)
                except Exception as e2:
                    errores.append(f"openpyxl manual: {str(e2)[:100]}")
        elif extension == '.xls':
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='xlrd', header=None)
            except Exception as e:
                errores.append(f"xlrd: {str(e)[:100]}")
        else:
            # Para otras extensiones (.xlsb, .xltx, .xltm)
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='openpyxl', header=None)
            except Exception as e:
                errores.append(f"openpyxl: {str(e)[:100]}")

        if not df_dict:
            logger.warning(f"No se pudo leer Excel {nombre_archivo}: {'; '.join(errores)}")
            raise HTTPException(
                status_code=400,
                detail=f"Error al leer Excel: {'; '.join(errores)}"
            )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=0.3*inch,
            rightMargin=0.3*inch,
            topMargin=0.4*inch,
            bottomMargin=0.4*inch
        )

        elements = []
        styles = getSampleStyleSheet()

        estilo_titulo = ParagraphStyle(
            'TituloHoja',
            parent=styles['Heading1'],
            fontName=FUENTE_PRINCIPAL,
            fontSize=14,
            spaceAfter=12
        )

        estilo_celda = ParagraphStyle(
            'CeldaTabla',
            fontName=FUENTE_PRINCIPAL,
            fontSize=9,
            leading=11,
        )

        for sheet_name, df in df_dict.items():
            elements.append(Paragraph(f"Hoja: {sheet_name}", estilo_titulo))

            df_clean = limpiar_dataframe(df)

            if df_clean.empty:
                continue

            data = []
            for row in df_clean.values.tolist():
                fila_procesada = []
                for c in row:
                    texto = formatear_celda(c)
                    fila_procesada.append(Paragraph(texto, estilo_celda))
                data.append(fila_procesada)

            if not data:
                continue

            ancho_util = landscape(A4)[0] - 0.6*inch
            col_widths = calcular_anchos_columnas_mejorado(df_clean, ancho_util)

            table = Table(data, colWidths=col_widths, repeatRows=1)

            estilo_grid = [
                ('FONTNAME', (0, 0), (-1, -1), FUENTE_PRINCIPAL),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.93, 0.93, 0.93)),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]

            table.setStyle(TableStyle(estilo_grid))

            elements.append(table)
            elements.append(PageBreak())

        doc.build(elements)
        return buffer.getvalue()

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error al convertir Excel a PDF: {str(e)}"
        )


async def convertir_excel_a_pdf(excel_bytes: bytes, nombre_archivo: str) -> bytes:
    """Convierte un archivo Excel a PDF de forma asíncrona con timeout dinámico."""
    timeout = calcular_timeout_excel(len(excel_bytes))
    file_size_mb = len(excel_bytes) / (1024 * 1024)

    logger.info(f"Iniciando conversión Excel a PDF para {nombre_archivo} ({file_size_mb:.2f}MB, timeout={timeout}s)")
    start_time = time.time()

    loop = asyncio.get_event_loop()

    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                _convertir_excel_a_pdf_sync,
                excel_bytes,
                nombre_archivo
            ),
            timeout=timeout
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Conversión Excel completada para {nombre_archivo} en {elapsed_time:.2f}s")

        return result
    except asyncio.TimeoutError:
        logger.error(f"Conversión Excel excedió timeout para {nombre_archivo} ({timeout}s)")
        raise HTTPException(
            status_code=408,
            detail=f"Conversión Excel excedió el tiempo límite de {timeout}s"
        )


async def convertir_imagen_a_pdf(imagen_bytes: bytes, nombre_archivo: str) -> bytes:
    """Convierte una imagen a PDF de forma asíncrona."""
    file_size_mb = len(imagen_bytes) / (1024 * 1024)
    timeout = 60  # 60 segundos para imágenes

    logger.info(f"Iniciando conversión imagen a PDF para {nombre_archivo} ({file_size_mb:.2f}MB)")
    start_time = time.time()

    loop = asyncio.get_event_loop()

    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                _convertir_imagen_a_pdf_sync,
                imagen_bytes,
                nombre_archivo
            ),
            timeout=timeout
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Conversión imagen completada para {nombre_archivo} en {elapsed_time:.2f}s")

        return result
    except asyncio.TimeoutError:
        logger.error(f"Conversión imagen excedió timeout para {nombre_archivo} ({timeout}s)")
        raise HTTPException(
            status_code=408,
            detail=f"Conversión imagen excedió el tiempo límite de {timeout}s"
        )


async def verificar_modelo_entrenado(model_id: str) -> bool:
    """Verifica si un modelo custom está entrenado en Azure DI."""
    url = f"{AZURE_DI_ENDPOINT}/formrecognizer/documentModels/{model_id}?api-version={API_VERSION}"

    timeout_config = httpx.Timeout(
        connect=settings.TIMEOUT_CONNECT,
        read=settings.TIMEOUT_READ,
        write=settings.TIMEOUT_WRITE,
        pool=5.0
    )

    try:
        async with httpx.AsyncClient(timeout=timeout_config) as client:
            response = await client.get(url)
            return response.status_code == 200
    except Exception:
        return False


async def extraer_datos_con_modelo(pdf_bytes: bytes, model_id: str) -> Optional[Dict[str, Any]]:
    """
    Extrae datos estructurados de un PDF usando un modelo custom de Azure DI.

    Args:
        pdf_bytes: Bytes del PDF a analizar
        model_id: ID del modelo custom a usar

    Returns:
        Diccionario con los campos extraídos, o None si hubo error
    """
    # Los modelos custom están en el contenedor azure-di-custom
    url = f"{AZURE_DI_ENDPOINT}/formrecognizer/documentModels/{model_id}:analyze?api-version=2023-07-31"

    headers = {
        "Content-Type": "application/pdf"
    }

    timeout_config = httpx.Timeout(
        connect=settings.TIMEOUT_CONNECT,
        read=300.0,  # 5 minutos para análisis
        write=settings.TIMEOUT_WRITE,
        pool=5.0
    )

    try:
        async with httpx.AsyncClient(timeout=timeout_config) as client:
            # Iniciar análisis
            print(f"[DEBUG] POST {url}")
            response = await client.post(url, headers=headers, content=pdf_bytes)
            print(f"[DEBUG] Respuesta inicial: {response.status_code}")

            if response.status_code != 202:
                print(f"[DEBUG] Error: esperaba 202, recibió {response.status_code}: {response.text}")
                return None

            # Obtener URL de operación
            operation_location = response.headers.get('Operation-Location')
            if not operation_location:
                print(f"[DEBUG] Error: no se recibió Operation-Location")
                return None

            print(f"[DEBUG] Polling en: {operation_location}")

            # Esperar a que termine el análisis (polling)
            max_intentos = 60  # 5 minutos máximo (60 * 5s)
            for intento in range(max_intentos):
                status_response = await client.get(operation_location)

                if status_response.status_code != 200:
                    print(f"[DEBUG] Error en polling: {status_response.status_code}")
                    return None

                resultado = status_response.json()
                status = resultado.get('status')
                print(f"[DEBUG] Intento {intento+1}: status={status}")

                if status == 'succeeded':
                    # Extraer datos del resultado
                    analyze_result = resultado.get('analyzeResult', {})
                    documentos = analyze_result.get('documents', [])
                    print(f"[DEBUG] Análisis exitoso. Documentos encontrados: {len(documentos)}")

                    if documentos:
                        # Retornar los campos del primer documento
                        fields = documentos[0].get('fields', {})
                        print(f"[DEBUG] Campos extraídos: {len(fields)}")

                        # Simplificar la estructura de los campos
                        datos_extraidos = {}
                        for field_name, field_data in fields.items():
                            if isinstance(field_data, dict):
                                # Extraer el valor según el tipo
                                if 'valueString' in field_data:
                                    datos_extraidos[field_name] = field_data['valueString']
                                elif 'valueNumber' in field_data:
                                    datos_extraidos[field_name] = field_data['valueNumber']
                                elif 'valueDate' in field_data:
                                    datos_extraidos[field_name] = field_data['valueDate']
                                elif 'valueArray' in field_data:
                                    datos_extraidos[field_name] = field_data['valueArray']
                                elif 'valueObject' in field_data:
                                    datos_extraidos[field_name] = field_data['valueObject']
                                elif 'content' in field_data:
                                    datos_extraidos[field_name] = field_data['content']

                        print(f"[DEBUG] Datos simplificados: {list(datos_extraidos.keys())}")
                        return datos_extraidos
                    return {}

                elif status == 'failed':
                    error_info = resultado.get('error', {})
                    print(f"[DEBUG] Análisis falló: {error_info}")
                    return None

                # Esperar 5 segundos antes de verificar de nuevo
                await asyncio.sleep(5)

            print(f"[DEBUG] Timeout: se excedió el máximo de intentos")
            return None

    except Exception as e:
        print(f"[DEBUG] Excepción en extraer_datos_con_modelo: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def _procesar_calidad_sync(pdf_bytes: bytes) -> tuple:
    """Función síncrona interna para procesamiento de calidad."""
    pdf_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    
    with suprimir_prints():
        resultados_paso_1 = paso_1_analizar_documento(pdf_doc)
        paginas_corregidas = paso_2_corregir_rotacion(pdf_doc, resultados_paso_1)
    
    if paginas_corregidas > 0:
        pdf_bytes = pdf_doc.tobytes()
    
    pdf_doc.close()
    return pdf_bytes, resultados_paso_1


async def clasificar_pdf_completo(pdf_bytes: bytes, nombre_archivo: str) -> Dict[str, Any]:
    """
    Flujo de clasificación de PDF (sin extracción de datos):
    1. Análisis y corrección de calidad
    2. Clasificación de páginas
    3. Segmentación de documentos

    Retorna documentos finales segmentados con alertas de calidad pero sin extracción de datos.
    """
    try:
        # Obtener número de páginas para timeout adaptativo
        pdf_doc_temp = fitz.open(stream=pdf_bytes, filetype="pdf")
        num_paginas = len(pdf_doc_temp)
        pdf_doc_temp.close()

        timeout_calidad = calcular_timeout_calidad(num_paginas)

        logger.info(f"Iniciando procesamiento de calidad para {nombre_archivo} ({num_paginas} páginas, timeout={timeout_calidad}s)")
        start_time = time.time()

        loop = asyncio.get_event_loop()
        pdf_bytes, resultados_paso_1 = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                _procesar_calidad_sync,
                pdf_bytes
            ),
            timeout=timeout_calidad
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Procesamiento de calidad completado para {nombre_archivo} en {elapsed_time:.2f}s")

        clasificaciones = await clasificar_documento_completo(pdf_bytes)

        logger.info(f"Iniciando segmentación para {nombre_archivo}")
        start_time = time.time()

        documentos_segmentados = await loop.run_in_executor(
            executor,
            segmentar_pdf,
            pdf_bytes,
            clasificaciones
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Segmentación completada para {nombre_archivo} en {elapsed_time:.2f}s - {len(documentos_segmentados)} documentos")

        alertas_por_documento = {}
        for resultado in resultados_paso_1:
            alertas = []

            if resultado['escaneada']:
                if 'INCLINADA' in resultado['orientacion']:
                    alertas.append({
                        "pagina": resultado['pagina'],
                        "tipo": "inclinado",
                        "descripcion": f"Página escaneada {resultado['orientacion']}"
                    })
                elif resultado['orientacion'] not in ['NORMAL', 'SIN TEXTO', 'SIN IMAGEN']:
                    alertas.append({
                        "pagina": resultado['pagina'],
                        "tipo": "escaneado",
                        "descripcion": f"Página escaneada: {resultado['orientacion']}"
                    })

            if resultado['rotacion_formal'] != 0:
                alertas.append({
                    "pagina": resultado['pagina'],
                    "tipo": "rotado",
                    "descripcion": f"Rotación de {resultado['rotacion_formal']}° corregida"
                })

            if not resultado['escaneada'] and resultado['orientacion'] == 'ROTADA':
                alertas.append({
                    "pagina": resultado['pagina'],
                    "tipo": "rotado",
                    "descripcion": "Texto vertical corregido"
                })

            if alertas:
                alertas_por_documento[resultado['pagina']] = alertas

        documentos_finales = []
        for idx, doc_seg in enumerate(documentos_segmentados):
            alertas_segmento = []
            for pagina in doc_seg['paginas']:
                if pagina in alertas_por_documento:
                    alertas_segmento.extend(alertas_por_documento[pagina])

            nombre_salida = f"{nombre_archivo.replace('.pdf', '')}_{doc_seg['tipo']}_{idx+1}.pdf"

            documentos_finales.append({
                "archivo_origen": nombre_archivo,
                "nombre_salida": nombre_salida,
                "tipo": doc_seg['tipo'],
                "paginas": doc_seg['paginas'],
                "pdf_bytes": doc_seg['pdf_bytes'],
                "alertas": alertas_segmento if alertas_segmento else None,
                "datos_extraidos": None
            })

        return {
            "documentos_finales": documentos_finales,
            "clasificaciones": clasificaciones,
            "error": None
        }

    except asyncio.TimeoutError:
        logger.error(f"Timeout en procesamiento de calidad para {nombre_archivo}")
        return {
            "documentos_finales": [],
            "clasificaciones": [],
            "error": f"Timeout en procesamiento de calidad ({timeout_calidad}s)"
        }
    except Exception as e:
        logger.error(f"Error en clasificar_pdf_completo para {nombre_archivo}: {str(e)}")
        return {
            "documentos_finales": [],
            "clasificaciones": [],
            "error": str(e)
        }


async def procesar_pdf_completo(pdf_bytes: bytes, nombre_archivo: str) -> Dict[str, Any]:
    """
    Flujo completo de procesamiento de PDF:
    1. Análisis y corrección de calidad
    2. Clasificación de páginas
    3. Segmentación de documentos
    4. Extracción de datos con modelos custom (si están entrenados)

    Retorna documentos finales segmentados con alertas de calidad y datos extraídos.
    """
    try:
        # Obtener número de páginas para timeout adaptativo
        pdf_doc_temp = fitz.open(stream=pdf_bytes, filetype="pdf")
        num_paginas = len(pdf_doc_temp)
        pdf_doc_temp.close()

        timeout_calidad = calcular_timeout_calidad(num_paginas)

        logger.info(f"Iniciando procesamiento de calidad para {nombre_archivo} ({num_paginas} páginas, timeout={timeout_calidad}s)")
        start_time = time.time()

        loop = asyncio.get_event_loop()
        pdf_bytes, resultados_paso_1 = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                _procesar_calidad_sync,
                pdf_bytes
            ),
            timeout=timeout_calidad
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Procesamiento de calidad completado para {nombre_archivo} en {elapsed_time:.2f}s")

        clasificaciones = await clasificar_documento_completo(pdf_bytes)

        logger.info(f"Iniciando segmentación para {nombre_archivo}")
        start_time = time.time()

        documentos_segmentados = await loop.run_in_executor(
            executor,
            segmentar_pdf,
            pdf_bytes,
            clasificaciones
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Segmentación completada para {nombre_archivo} en {elapsed_time:.2f}s - {len(documentos_segmentados)} documentos")

        alertas_por_documento = {}
        for resultado in resultados_paso_1:
            alertas = []

            if resultado['escaneada']:
                if 'INCLINADA' in resultado['orientacion']:
                    alertas.append({
                        "pagina": resultado['pagina'],
                        "tipo": "inclinado",
                        "descripcion": f"Página escaneada {resultado['orientacion']}"
                    })
                elif resultado['orientacion'] not in ['NORMAL', 'SIN TEXTO', 'SIN IMAGEN']:
                    alertas.append({
                        "pagina": resultado['pagina'],
                        "tipo": "escaneado",
                        "descripcion": f"Página escaneada: {resultado['orientacion']}"
                    })

            if resultado['rotacion_formal'] != 0:
                alertas.append({
                    "pagina": resultado['pagina'],
                    "tipo": "rotado",
                    "descripcion": f"Rotación de {resultado['rotacion_formal']}° corregida"
                })

            if not resultado['escaneada'] and resultado['orientacion'] == 'ROTADA':
                alertas.append({
                    "pagina": resultado['pagina'],
                    "tipo": "rotado",
                    "descripcion": "Texto vertical corregido"
                })

            if alertas:
                alertas_por_documento[resultado['pagina']] = alertas

        documentos_finales = []
        for idx, doc_seg in enumerate(documentos_segmentados):
            alertas_segmento = []
            for pagina in doc_seg['paginas']:
                if pagina in alertas_por_documento:
                    alertas_segmento.extend(alertas_por_documento[pagina])

            nombre_salida = f"{nombre_archivo.replace('.pdf', '')}_{doc_seg['tipo']}_{idx+1}.pdf"

            # Obtener el model_id correspondiente al tipo de documento
            tipo_documento = doc_seg['tipo']
            model_id = DOCUMENT_TYPE_TO_MODEL.get(tipo_documento)

            # Inicializar datos extraídos como None
            datos_extraidos = None

            # Si hay un modelo asociado, intentar extraer datos
            if model_id:
                try:
                    # Verificar si el modelo está entrenado
                    modelo_entrenado = await verificar_modelo_entrenado(model_id)
                    print(f"[DEBUG] Tipo: {tipo_documento}, Model ID: {model_id}, Entrenado: {modelo_entrenado}")

                    if modelo_entrenado:
                        # Extraer datos usando el modelo custom
                        print(f"[DEBUG] Iniciando extracción para {model_id}...")
                        datos_extraidos = await extraer_datos_con_modelo(
                            doc_seg['pdf_bytes'],
                            model_id
                        )
                        print(f"[DEBUG] Extracción completada. Datos: {datos_extraidos is not None}")
                except Exception as e:
                    # Si hay error en la extracción, continuar sin datos extraídos
                    print(f"[DEBUG] Error en extracción: {str(e)}")
                    datos_extraidos = None

            documentos_finales.append({
                "archivo_origen": nombre_archivo,
                "nombre_salida": nombre_salida,
                "tipo": doc_seg['tipo'],
                "paginas": doc_seg['paginas'],
                "pdf_bytes": doc_seg['pdf_bytes'],
                "alertas": alertas_segmento if alertas_segmento else None,
                "datos_extraidos": datos_extraidos
            })

        return {
            "documentos_finales": documentos_finales,
            "clasificaciones": clasificaciones,
            "error": None
        }

    except asyncio.TimeoutError:
        logger.error(f"Timeout en procesamiento de calidad para {nombre_archivo}")
        return {
            "documentos_finales": [],
            "clasificaciones": [],
            "error": f"Timeout en procesamiento de calidad ({timeout_calidad}s)"
        }
    except Exception as e:
        logger.error(f"Error en procesar_pdf_completo para {nombre_archivo}: {str(e)}")
        return {
            "documentos_finales": [],
            "clasificaciones": [],
            "error": str(e)
        }


@sgd_router.get("/consultar/{codigo_despacho}")
async def consultar_despacho(codigo_despacho: str, token: str = Depends(verify_api_token)):
    """
    Consulta la información del despacho y lista los documentos disponibles.
    Soporta tanto ID interno como código visible.
    """
    if not settings.BEARER_TOKEN:
        raise HTTPException(status_code=500, detail="BEARER_TOKEN no configurado")
    
    datos_despacho_detalle = await consultar_despacho_detalle(codigo_despacho, settings.BEARER_TOKEN)
    
    if datos_despacho_detalle and "data" in datos_despacho_detalle:
        despacho_data = datos_despacho_detalle["data"]
        
        if isinstance(despacho_data, dict):
            codigo_visible = str(despacho_data.get("codigo", "N/A"))
            cliente = despacho_data.get("cliente", {})
            cliente_nombre = cliente.get("nombre", "N/A") if isinstance(cliente, dict) else "N/A"
            
            documentos_list = despacho_data.get("documentos", [])
            if not isinstance(documentos_list, list):
                documentos_list = []
            
            documentos_simplificados = []
            for doc in documentos_list:
                if isinstance(doc, dict):
                    tipo = doc.get("tipo", {})
                    documentos_simplificados.append({
                        "nombre": tipo.get("nombre", "Sin nombre") if isinstance(tipo, dict) else "Sin nombre",
                        "estado": doc.get("estado", "N/A"),
                        "fecha_recepcion": doc.get("fecha_recepcion", "N/A")
                    })
            
            usuarios_list = despacho_data.get("usuarios", [])
            if not isinstance(usuarios_list, list):
                usuarios_list = []
            
            pedidores = []
            jefes_operaciones = []
            
            for user in usuarios_list:
                if isinstance(user, dict):
                    role_name = user.get("role_name", "")
                    nombre = user.get("name", "")
                    
                    if role_name in ("pedidor", "pedidor_exportaciones"):
                        pedidores.append(nombre)
                    elif role_name == "jefe_operaciones":
                        jefes_operaciones.append(nombre)
            
            usuarios_datos = {
                "pedidor": pedidores if pedidores else None,
                "jefe_operaciones": jefes_operaciones if jefes_operaciones else None
            }
            
            return {
                "codigo_despacho": codigo_visible,
                "id_interno": str(despacho_data.get("id", "N/A")),
                "cliente": cliente_nombre,
                "estado": despacho_data.get("estado_despacho", "N/A"),
                "tipo": despacho_data.get("tipo_despacho", "N/A"),
                "total_documentos": len(documentos_simplificados),
                "documentos": documentos_simplificados,
                "usuarios": usuarios_datos
            }
    
    documentos_base64_list = await consultar_documentacion(codigo_despacho, settings.BEARER_TOKEN)
    
    if documentos_base64_list and isinstance(documentos_base64_list, list):
        documentos_info = []
        for doc in documentos_base64_list:
            if isinstance(doc, dict):
                documentos_info.append({
                    "nombre": doc.get("nombre_documento", "Sin nombre"),
                    "documento_id": doc.get("documento_id", "N/A")
                })
        
        return {
            "codigo": codigo_despacho,
            "mensaje": "Información limitada: solo se encontraron documentos",
            "total_documentos": len(documentos_info),
            "documentos": documentos_info
        }
    
    raise HTTPException(status_code=404, detail="Despacho no encontrado")


@sgd_router.post("/clasificar/{codigo_despacho}", response_model=ProcesamientoResponse)
async def clasificar_despacho(codigo_despacho: str, token: str = Depends(verify_api_token)):
    """
    Clasifica los documentos del despacho:
    1. Análisis y corrección de calidad
    2. Clasificación de páginas
    3. Segmentación de documentos

    No incluye extracción de datos con modelos.
    Almacena solo documentos finales segmentados en memoria.
    """
    if not settings.BEARER_TOKEN:
        raise HTTPException(status_code=500, detail="BEARER_TOKEN no configurado")

    datos_despacho_detalle = await consultar_despacho_detalle(codigo_despacho, settings.BEARER_TOKEN)

    codigo_visible = None
    cliente_nombre = "N/A"
    estado_despacho = "N/A"
    tipo_despacho = "N/A"

    if datos_despacho_detalle and "data" in datos_despacho_detalle:
        despacho_data = datos_despacho_detalle["data"]
        if isinstance(despacho_data, dict):
            codigo_visible = str(despacho_data.get("codigo", ""))
            cliente = despacho_data.get("cliente", {})
            cliente_nombre = cliente.get("nombre", "N/A") if isinstance(cliente, dict) else "N/A"
            estado_despacho = despacho_data.get("estado_despacho", "N/A")
            tipo_despacho = despacho_data.get("tipo_despacho", "N/A")

    documentos_base64_list = None

    if codigo_visible:
        documentos_base64_list = await consultar_documentacion(codigo_visible, settings.BEARER_TOKEN)

    if not documentos_base64_list:
        documentos_base64_list = await consultar_documentacion(codigo_despacho, settings.BEARER_TOKEN)

    if not documentos_base64_list or not isinstance(documentos_base64_list, list):
        raise HTTPException(status_code=404, detail="No se pudieron obtener los documentos")

    todos_documentos_finales = []

    for doc in documentos_base64_list:
        if isinstance(doc, dict):
            base64_data = doc.get("documento", "")

            if ',' in base64_data:
                _, base64_content = base64_data.split(',', 1)
            else:
                base64_content = base64_data

            try:
                file_bytes = base64.b64decode(base64_content)
                nombre_documento = doc.get("nombre_documento", "documento.pdf")

                validar_tamano_archivo(file_bytes)

                if es_archivo_excel(nombre_documento):
                    if not validar_excel(file_bytes, nombre_documento):
                        continue
                    try:
                        pdf_bytes = await convertir_excel_a_pdf(file_bytes, nombre_documento)
                        nombre_procesamiento = nombre_documento.rsplit('.', 1)[0] + '.pdf'
                    except Exception:
                        continue
                elif es_archivo_imagen(nombre_documento):
                    if not validar_imagen(file_bytes, nombre_documento):
                        continue
                    try:
                        pdf_bytes = await convertir_imagen_a_pdf(file_bytes, nombre_documento)
                        nombre_procesamiento = nombre_documento.rsplit('.', 1)[0] + '.pdf'
                    except Exception:
                        continue
                else:
                    if not validar_pdf(file_bytes):
                        continue
                    pdf_bytes = file_bytes
                    nombre_procesamiento = nombre_documento

                resultado = await clasificar_pdf_completo(pdf_bytes, nombre_procesamiento)

                if resultado["error"]:
                    continue

                todos_documentos_finales.extend(resultado["documentos_finales"])

            except Exception:
                continue

    documentos_finales_sgd[codigo_despacho] = todos_documentos_finales

    docs_response = []
    for doc_final in todos_documentos_finales:
        docs_response.append(DocumentoFinal(
            archivo_origen=doc_final["archivo_origen"],
            nombre_salida=doc_final["nombre_salida"],
            tipo=doc_final["tipo"],
            paginas=doc_final["paginas"],
            alertas=[Alerta(**alerta) for alerta in doc_final["alertas"]] if doc_final["alertas"] else None,
            datos_extraidos=doc_final.get("datos_extraidos")
        ))

    return ProcesamientoResponse(
        codigo_despacho=codigo_despacho,
        cliente=cliente_nombre,
        estado=estado_despacho,
        tipo=tipo_despacho,
        total_documentos_segmentados=len(docs_response),
        documentos=docs_response
    )


@sgd_router.post("/procesar/{codigo_despacho}", response_model=ProcesamientoResponse)
async def procesar_despacho(codigo_despacho: str, token: str = Depends(verify_api_token)):
    """
    Procesa el despacho completo:
    1. Análisis y corrección de calidad
    2. Clasificación de páginas
    3. Segmentación de documentos

    Almacena solo documentos finales segmentados en memoria.
    """
    if not settings.BEARER_TOKEN:
        raise HTTPException(status_code=500, detail="BEARER_TOKEN no configurado")
    
    datos_despacho_detalle = await consultar_despacho_detalle(codigo_despacho, settings.BEARER_TOKEN)
    
    codigo_visible = None
    cliente_nombre = "N/A"
    estado_despacho = "N/A"
    tipo_despacho = "N/A"
    
    if datos_despacho_detalle and "data" in datos_despacho_detalle:
        despacho_data = datos_despacho_detalle["data"]
        if isinstance(despacho_data, dict):
            codigo_visible = str(despacho_data.get("codigo", ""))
            cliente = despacho_data.get("cliente", {})
            cliente_nombre = cliente.get("nombre", "N/A") if isinstance(cliente, dict) else "N/A"
            estado_despacho = despacho_data.get("estado_despacho", "N/A")
            tipo_despacho = despacho_data.get("tipo_despacho", "N/A")
    
    documentos_base64_list = None
    
    if codigo_visible:
        documentos_base64_list = await consultar_documentacion(codigo_visible, settings.BEARER_TOKEN)
    
    if not documentos_base64_list:
        documentos_base64_list = await consultar_documentacion(codigo_despacho, settings.BEARER_TOKEN)
    
    if not documentos_base64_list or not isinstance(documentos_base64_list, list):
        raise HTTPException(status_code=404, detail="No se pudieron obtener los documentos")
    
    todos_documentos_finales = []
    
    for doc in documentos_base64_list:
        if isinstance(doc, dict):
            base64_data = doc.get("documento", "")
            
            if ',' in base64_data:
                _, base64_content = base64_data.split(',', 1)
            else:
                base64_content = base64_data
            
            try:
                file_bytes = base64.b64decode(base64_content)
                nombre_documento = doc.get("nombre_documento", "documento.pdf")
                
                validar_tamano_archivo(file_bytes)

                if es_archivo_excel(nombre_documento):
                    if not validar_excel(file_bytes, nombre_documento):
                        continue
                    try:
                        pdf_bytes = await convertir_excel_a_pdf(file_bytes, nombre_documento)
                        nombre_procesamiento = nombre_documento.rsplit('.', 1)[0] + '.pdf'
                    except Exception:
                        continue
                elif es_archivo_imagen(nombre_documento):
                    if not validar_imagen(file_bytes, nombre_documento):
                        continue
                    try:
                        pdf_bytes = await convertir_imagen_a_pdf(file_bytes, nombre_documento)
                        nombre_procesamiento = nombre_documento.rsplit('.', 1)[0] + '.pdf'
                    except Exception:
                        continue
                else:
                    if not validar_pdf(file_bytes):
                        continue
                    pdf_bytes = file_bytes
                    nombre_procesamiento = nombre_documento

                resultado = await procesar_pdf_completo(pdf_bytes, nombre_procesamiento)
                
                if resultado["error"]:
                    continue
                
                todos_documentos_finales.extend(resultado["documentos_finales"])
                
            except Exception:
                continue
    
    documentos_finales_sgd[codigo_despacho] = todos_documentos_finales
    
    docs_response = []
    for doc_final in todos_documentos_finales:
        docs_response.append(DocumentoFinal(
            archivo_origen=doc_final["archivo_origen"],
            nombre_salida=doc_final["nombre_salida"],
            tipo=doc_final["tipo"],
            paginas=doc_final["paginas"],
            alertas=[Alerta(**alerta) for alerta in doc_final["alertas"]] if doc_final["alertas"] else None,
            datos_extraidos=doc_final.get("datos_extraidos")
        ))

    return ProcesamientoResponse(
        codigo_despacho=codigo_despacho,
        cliente=cliente_nombre,
        estado=estado_despacho,
        tipo=tipo_despacho,
        total_documentos_segmentados=len(docs_response),
        documentos=docs_response
    )


@documentos_router.post("/clasificar", response_model=ProcesamientoIndividualResponse)
async def clasificar_documento_individual(file: UploadFile = File(...), token: str = Depends(verify_api_token)):
    """
    Clasifica un documento individual:
    - Soporta archivos PDF, Excel (.xls, .xlsx, .xlsm, etc.) e imágenes (.jpg, .png, etc.)
    - Los archivos Excel e imágenes se convierten automáticamente a PDF
    - Valida el tipo MIME real del archivo

    Flujo:
    1. Validación de tipo de archivo
    2. Conversión Excel/imagen a PDF (si aplica)
    3. Análisis y corrección de calidad
    4. Clasificación de páginas
    5. Segmentación de documentos

    No incluye extracción de datos con modelos.
    Almacena solo documentos finales segmentados en memoria.
    """
    nombre_archivo = file.filename.lower()
    es_pdf = nombre_archivo.endswith('.pdf')
    es_excel = es_archivo_excel(file.filename)
    es_imagen = es_archivo_imagen(file.filename)

    if not es_pdf and not es_excel and not es_imagen:
        raise HTTPException(
            status_code=400,
            detail="Solo se aceptan archivos PDF, Excel (.xls, .xlsx, .xlsm, .xlsb, .xltx, .xltm) o imágenes (.jpg, .jpeg, .png, .gif, .bmp, .tiff, .webp)"
        )

    try:
        file_bytes = await file.read()

        validar_tamano_archivo(file_bytes)

        if es_excel:
            if not validar_excel(file_bytes, file.filename):
                raise HTTPException(
                    status_code=400,
                    detail="El archivo no es un Excel válido"
                )
            pdf_bytes = await convertir_excel_a_pdf(file_bytes, file.filename)
            nombre_procesamiento = file.filename.rsplit('.', 1)[0] + '.pdf'
        elif es_imagen:
            if not validar_imagen(file_bytes, file.filename):
                raise HTTPException(
                    status_code=400,
                    detail="El archivo no es una imagen válida"
                )
            pdf_bytes = await convertir_imagen_a_pdf(file_bytes, file.filename)
            nombre_procesamiento = file.filename.rsplit('.', 1)[0] + '.pdf'
        else:
            if not validar_pdf(file_bytes):
                raise HTTPException(
                    status_code=400,
                    detail="El archivo no es un PDF válido"
                )
            pdf_bytes = file_bytes
            nombre_procesamiento = file.filename

        resultado = await clasificar_pdf_completo(pdf_bytes, nombre_procesamiento)

        if resultado["error"]:
            raise HTTPException(status_code=500, detail=f"Error al clasificar: {resultado['error']}")

        documentos_finales_individuales[file.filename] = resultado["documentos_finales"]

        docs_response = []
        for doc_final in resultado["documentos_finales"]:
            docs_response.append(DocumentoFinal(
                archivo_origen=doc_final["archivo_origen"],
                nombre_salida=doc_final["nombre_salida"],
                tipo=doc_final["tipo"],
                paginas=doc_final["paginas"],
                alertas=[Alerta(**alerta) for alerta in doc_final["alertas"]] if doc_final["alertas"] else None,
                datos_extraidos=doc_final.get("datos_extraidos")
            ))

        return ProcesamientoIndividualResponse(
            archivo_origen=file.filename,
            total_documentos_segmentados=len(docs_response),
            documentos=docs_response
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al clasificar documento: {str(e)}")


@documentos_router.post("/procesar", response_model=ProcesamientoIndividualResponse)
async def procesar_documento_individual(file: UploadFile = File(...), token: str = Depends(verify_api_token)):
    """
    Procesa un documento individual completo:
    - Soporta archivos PDF, Excel (.xls, .xlsx, .xlsm, etc.) e imágenes (.jpg, .png, etc.)
    - Los archivos Excel e imágenes se convierten automáticamente a PDF
    - Valida el tipo MIME real del archivo

    Flujo:
    1. Validación de tipo de archivo
    2. Conversión Excel/imagen a PDF (si aplica)
    3. Análisis y corrección de calidad
    4. Clasificación de páginas
    5. Segmentación de documentos

    Almacena solo documentos finales segmentados en memoria.
    """
    nombre_archivo = file.filename.lower()
    es_pdf = nombre_archivo.endswith('.pdf')
    es_excel = es_archivo_excel(file.filename)
    es_imagen = es_archivo_imagen(file.filename)

    if not es_pdf and not es_excel and not es_imagen:
        raise HTTPException(
            status_code=400,
            detail="Solo se aceptan archivos PDF, Excel (.xls, .xlsx, .xlsm, .xlsb, .xltx, .xltm) o imágenes (.jpg, .jpeg, .png, .gif, .bmp, .tiff, .webp)"
        )

    try:
        file_bytes = await file.read()

        validar_tamano_archivo(file_bytes)

        if es_excel:
            if not validar_excel(file_bytes, file.filename):
                raise HTTPException(
                    status_code=400,
                    detail="El archivo no es un Excel válido"
                )
            pdf_bytes = await convertir_excel_a_pdf(file_bytes, file.filename)
            nombre_procesamiento = file.filename.rsplit('.', 1)[0] + '.pdf'
        elif es_imagen:
            if not validar_imagen(file_bytes, file.filename):
                raise HTTPException(
                    status_code=400,
                    detail="El archivo no es una imagen válida"
                )
            pdf_bytes = await convertir_imagen_a_pdf(file_bytes, file.filename)
            nombre_procesamiento = file.filename.rsplit('.', 1)[0] + '.pdf'
        else:
            if not validar_pdf(file_bytes):
                raise HTTPException(
                    status_code=400,
                    detail="El archivo no es un PDF válido"
                )
            pdf_bytes = file_bytes
            nombre_procesamiento = file.filename

        resultado = await procesar_pdf_completo(pdf_bytes, nombre_procesamiento)
        
        if resultado["error"]:
            raise HTTPException(status_code=500, detail=f"Error al procesar: {resultado['error']}")
        
        documentos_finales_individuales[file.filename] = resultado["documentos_finales"]
        
        docs_response = []
        for doc_final in resultado["documentos_finales"]:
            docs_response.append(DocumentoFinal(
                archivo_origen=doc_final["archivo_origen"],
                nombre_salida=doc_final["nombre_salida"],
                tipo=doc_final["tipo"],
                paginas=doc_final["paginas"],
                alertas=[Alerta(**alerta) for alerta in doc_final["alertas"]] if doc_final["alertas"] else None,
                datos_extraidos=doc_final.get("datos_extraidos")
            ))

        return ProcesamientoIndividualResponse(
            archivo_origen=file.filename,
            total_documentos_segmentados=len(docs_response),
            documentos=docs_response
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al procesar documento: {str(e)}")


# ============================================
# ADMIN ENDPOINTS - Gestión de Tokens
# ============================================

@admin_router.get("/", response_model=List[TokenInfo])
async def listar_tokens(admin_token: str = Depends(verify_admin_token)):
    """
    Lista todos los tokens de la API con su metadata.

    **Requiere token de administrador.**

    - Muestra tokens enmascarados por seguridad
    - Incluye fecha de creación y último uso
    - Muestra estado activo/inactivo
    - Incluye tokens del gestor (tokens.json) y tokens legacy (.env)
    """
    # Tokens del gestor (tokens.json)
    tokens = token_manager.list_tokens()

    # Agregar tokens del .env (API_TOKENS)
    env_tokens = get_valid_api_tokens()
    for idx, token_value in enumerate(env_tokens, start=1):
        # Enmascarar token
        masked_token = f"{token_value[:8]}...{token_value[-4:]}" if len(token_value) > 12 else "***"

        tokens.append({
            "id": f"env-token-{idx}",
            "name": f"Token de .env #{idx}",
            "masked_token": masked_token,
            "created_at": "N/A",
            "created_by": "env-config",
            "last_used": None,
            "is_active": True
        })

    return tokens


@admin_router.post("/generate", response_model=TokenCreateResponse)
async def generar_token(
    request: TokenCreateRequest,
    admin_token: str = Depends(verify_admin_token)
):
    """
    Genera un nuevo token de autenticación API.

    **Requiere token de administrador.**

    - **name**: Nombre descriptivo del token o usuario

    **IMPORTANTE**: El token completo solo se muestra una vez.
    Guárdalo en un lugar seguro.
    """
    result = token_manager.generate_token(
        name=request.name,
        created_by="admin"
    )
    return result


@admin_router.delete("/{token_id}", response_model=TokenDeleteResponse)
async def eliminar_token(
    token_id: str,
    admin_token: str = Depends(verify_admin_token)
):
    """
    Elimina un token por su ID.

    **Requiere token de administrador.**

    - **token_id**: ID del token a eliminar (obtenido desde listar_tokens)

    El token eliminado dejará de funcionar inmediatamente.
    """
    success = token_manager.delete_token(token_id)

    if success:
        return TokenDeleteResponse(
            success=True,
            message=f"Token {token_id} eliminado exitosamente"
        )
    else:
        raise HTTPException(
            status_code=404,
            detail=f"Token {token_id} no encontrado"
        )


app.include_router(sgd_router)
app.include_router(documentos_router)
app.include_router(modelos_router)
app.include_router(admin_router)


@app.get("/")
async def root():
    return {"status": "online", "service": "API Docs"}